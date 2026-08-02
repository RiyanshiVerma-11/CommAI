import sqlite3
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_excel(db_path: str, output_path: str):
    """
    Connects to the SQLite database and exports all tables into a single formatted Excel workbook.
    """
    if not os.path.exists(db_path):
        print(f"Error: Database file does not exist at '{db_path}'")
        return False

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Query all tables in the database
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = [row[0] for row in cursor.fetchall() if not row[0].startswith("sqlite_")]

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    if wb.active:
        wb.remove(wb.active)

    # Style definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    cell_font = Font(name=font_family, size=10, color="333333")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_side = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for table in tables:
        ws = wb.create_sheet(title=table.capitalize())
        
        # Get table columns
        cursor.execute(f"PRAGMA table_info({table});")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Query all records
        cursor.execute(f"SELECT * FROM {table};")
        rows = cursor.fetchall()

        # Add headers
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = cell_border
        
        # Add rows with zebra styling
        for row_idx, row in enumerate(rows, start=2):
            ws.append(row)
            fill_to_apply = zebra_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = cell_border
                cell.alignment = cell_alignment
                if fill_to_apply:
                    cell.fill = fill_to_apply

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # If cell is in row 1, add extra buffer
                val_len = len(val) + 4 if cell.row == 1 else len(val)
                if val_len > max_len:
                    max_len = val_len
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len, 10), 45)

        print(f"Exported table '{table}' with {len(rows)} records.")

    wb.save(output_path)
    conn.close()
    print(f"Database successfully exported to '{output_path}'")
    return True

if __name__ == "__main__":
    # If run standalone, detect DB path
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(base_dir, "comm_platform.db")
    out = os.path.join(base_dir, "database_export.xlsx")
    export_excel(db, out)
